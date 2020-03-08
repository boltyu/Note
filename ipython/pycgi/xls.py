import cgi
import cgitb
import os,time,datetime
import openpyxl
xls_name_pos = 2
xls_temp_pos = 4
xls_date_pos = 2
xls_date_data = "班级：      小一         日期:" + datetime.date.today().strftime("%Y年%m月%d日") +"                记录人： 张若媛"

tips = ""
form = cgi.FieldStorage()
yname = form.getvalue("yname"," ")
ytemp = form.getvalue("ytemp"," ")
ypos = form.getvalue("ypos"," ")
ytime = form.getvalue("ytime"," ")
ytrans = form.getvalue("ytrans"," ")
ifaccs = form.getvalue("ifaccs"," ")
nowdate = datetime.date.today().strftime("%Y年%m月%d日")
yname = "高明允"
namelist = ["bias0","bias1","bias2","bias3","bias4","张洺瑀","何昊轩","池芸晗","王甯皓","胜楷恩","李靖涵","刘佳祎","张彩嘉","张嘉凝","田意和","祖沛豪","刘子晨","俞心明","唐睿泽","陈诗玥","王玥涵","张隽芃","张稷中","马子昕","柴鸿飞","苏则平","苏矞媛","刘瑛童","李艾佳","李梓源","曲宜洋","高明允"]
try:
    yindex = namelist.index(yname)
    xls_name_pos = 2
    xls_temp_pos = 4
    xls_date_pos = 2
    xls_date_data = "班级：      小一         日期:" + datetime.date.today().strftime("%Y年%m月%d日") +"                记录人： 张若媛"
    xls_path = "ipython/pyandhtml/"
    xlsfilepath= xls_path + "小一班健康情况记录"+nowdate+".xlsx"
    workbook = ""
    if(not os.path.exists(xlsfilepath)):
        workbook=openpyxl.load_workbook("ipython/pyandhtml/base.xlsx")
    else:
        workbook=openpyxl.load_workbook(xlsfilepath)
    worksheet = workbook["Sheet1"]
    worksheet.cell(5,xls_temp_pos).value = "完美"
    bd = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style="thin",color="000000"),
                                top=openpyxl.styles.Side(border_style="thin",color="000000"),
                                right=openpyxl.styles.Side(border_style="thin",color="000000"),
                                bottom=openpyxl.styles.Side(border_style="thin",color="000000"))
    worksheet.cell(5,xls_temp_pos).border = bd
    
    worksheet.border = bd
    workbook.save(xlsfilepath)
except ValueError as e:
    print("孩子" + yname + "不在本班级中，请检查以上信息")

    