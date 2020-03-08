#!/usr/bin/env python3

import cgi
import cgitb
import os,time,datetime
import openpyxl

cgitb.enable()

tips = ""
form = cgi.FieldStorage()
yname = form.getvalue("yname"," ")
ytemp = form.getvalue("ytemp"," ")
ymyheal = form.getvalue("ymyheal"," ")
yfmheal = form.getvalue("yfmheal"," ")
ypos = form.getvalue("ypos"," ")
ytimem = form.getvalue("ytimem"," ")
ytimed = form.getvalue("ytimed"," ")
ytrans = form.getvalue("ytrans"," ")
ifaccs = form.getvalue("ifaccs"," ")
nowdate = datetime.date.today().strftime("%Y年%m月%d日")
bd = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style="thin",color="000000"),
                                top=openpyxl.styles.Side(border_style="thin",color="000000"),
                                right=openpyxl.styles.Side(border_style="thin",color="000000"),
                                bottom=openpyxl.styles.Side(border_style="thin",color="000000"))
workbook = ""
worksheet = ""

def fillcell(line,column,value):
    worksheet.cell(line,column).value = value
    worksheet.cell(line,column).border = bd

if(yname != " " and ytemp != " "):
    namelist = ["bias0","bias1","bias2","bias3","bias4","张洺瑀","何昊轩","池芸晗","王甯皓","胜楷恩","李靖涵","刘佳祎","张彩嘉","张嘉凝","田意和","祖沛豪","刘子晨","俞心明","唐睿泽","陈诗玥","王玥涵","张隽芃","张稷中","马子昕","柴鸿飞","苏则平","苏矞媛","刘瑛童","李艾佳","李梓源","曲宜洋","高明允"]
    try:
        yindex = namelist.index(yname)
        xls_pos_bias = 0
        if(yindex > 24):
            if(yindex > 27):
                xls_pos_bias = 10
            else:
                xls_pos_bias = 5
        xls_temp_pos = 4 + xls_pos_bias
        xls_myhealth_pos = xls_temp_pos + 1
        xls_familyhealth_pos = xls_temp_pos + 2
        xls_ifaccess_pos = xls_temp_pos + 3
        xls_familyhealth2_pos = xls_temp_pos + 4
        xls_date_pos = xls_temp_pos + 5
        xls_trans_pos = xls_temp_pos + 6
        xls_path = "../html/xls/"
        xlsfilepath= xls_path + "小一班健康情况记录"+nowdate+".xlsx"
        
        if(not os.path.exists(xlsfilepath)):
            workbook=openpyxl.load_workbook("/doc/base.xlsx")
            worksheet = workbook["Sheet1"]
            worksheet.cell(2,1).value = "班级：小一         日期:" + datetime.date.today().strftime("%Y年%m月%d日") +"                记录人： 张若媛"
        else:
            workbook=openpyxl.load_workbook(xlsfilepath)
            worksheet = workbook["Sheet1"]
        
        fillcell(yindex,xls_temp_pos,ytemp)
        fillcell(yindex,xls_myhealth_pos,ymyheal)
        fillcell(yindex,xls_familyhealth_pos,yfmheal)
        fillcell(yindex,xls_ifaccess_pos,ifaccs)
        if(yindex > 27):
            fillcell(yindex,xls_date_pos,ytimem+"."+ytimed)
            fillcell(yindex,xls_trans_pos,ytrans)
        fillcell(yindex,21,ypos)
        workbook.save(xlsfilepath)
        tips = tips + "  孩子" + yname + "今日信息已提交"
        '''print("""Content-type: text/html

<html>
<head>
<meta http-equiv="Content-Type" content="text/html; charset=utf-8" />
<meta name="viewport" content="width=device-width,height=device-height,inital-scale=1.0,maximum-scale=1.0,user-scalable=no;">
<title>小一班</title>
</head>
<h1>%s</h1>
</body>
</html> 
""" % tips)
        exit(0)'''
    except ValueError as ex:
        tips = "孩子" + yname + "不在本班级中，请检查以上信息"

print("""Content-type: text/html

<html>
<head>
<meta http-equiv="Content-Type" content="text/html; charset=utf-8" />
<meta name="viewport" content="width=device-width,height=device-height,inital-scale=1.0,maximum-scale=1.0,user-scalable=no;">
<title>小一班</title>
</head>

<body>
    <h1>%s</h1>
    <h1>小一班登记表</h1>
   <form action="t.html" method="post" id="user_form">
    幼儿姓名<input type="text" name="yname" size=12/><br/>
    今日体温<input type="text" name="ytemp" size=12/><br/>
    幼儿健康状况<select name="ymyheal">
        <option value="健康" selected="selected">健康</option>
        <option value="异常">异常</option>
        </select><br/>
    家人健康状况<select name="yfmheal">
        <option value="健康" selected="selected">健康</option>
        <option value="异常">异常</option>
        </select><br/>
    目前所在地<select name="ypos">
        <option value="北京" selected="selected">北京</option>
        <option value="福建">福建</option>
        <option value="湖北">湖北</option>
        <option value="哈尔滨">哈尔滨</option>
        <option value="其它">其它</option>
        </select><br/>
    返京时间<select name="ytimem">
            <option value="2" selected="selected">2月</option>
            <option value="3">3月</option>
            <option value="4">4月</option>
        </select>
        <input type="text" name="ytimed" size=3/>日<br/>
    返京交通工具<select name="ytrans">
        <option value="飞机" selected="selected">飞机</option>
        <option value="火车">火车</option>
        <option value="汽车">汽车</option>
        <option value="轮船">轮船</option>
        <option value="其它">其它</option>
        </select><br/>
    有无接触湖北人员或是接触患有新冠状病毒肺炎的人员<select name="ifaccs">
        <option value="否" selected="selected">否</option>
        <option value="是">是</option>
        </select><br/>
    <input type="submit" />
    <br/>
    <br/>
    %s
</form>
</body>
</html>
""" %(nowdate,tips))
